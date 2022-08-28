##############################################
# Estos son los imports para crear pivot tables
import win32com.client as win32
from pywintypes import com_error
from pathlib import Path
import sys
import numpy as np  # only used for synthetic data
import random  # only used for synthetic data
import pythoncom
import xlwings as xw
win32c = win32.constants
##############################################
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, date
from openpyxl import Workbook

# Esto libreria nos ayuda a crear pivot tables
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

##########################  IMPORT APP ##################################
# Es super necesario llamar al APP para poder utilizar las libreria propia
# Como cunado necesitamos llamar al directorio raiz de la applicacion a la hora de importar ficheros.
from run import app

######################
#Tablas e imagenes
# Primero importo algunas funciones para tablas e imagenes
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.chart import PieChart, Reference, Series,PieChart3D,LineChart, BarChart


class Format:

#---------- CREATE NEW COLUMN -MONTH-YEAR- --------------
    def new_column(filename):

        PATH = 'apps/uploads/tempData/'
        ##################
        #  OPEN FILE
        ##################
        # Ahora vamos a buscar el archivo y cargarlo
        wb = load_workbook(PATH + filename)
        ws = wb.active

        # Ahora creamos la cell que queremos agregar
        # Hay que tener en cuanta el tamaño del sheet para poder ubicarla bien en la columna correcta
        # Creo una variable donde guardo la cell

        newCol = ws['AE1']
        newCol.value = 'Month-Year'

        #################
        #  FIND SIZE DF
        #################
        # Podemos averiguar el tamaño de la coumna facilment con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(PATH + filename, engine='openpyxl')
        rows = df.shape[0]
        cols = df.shape[1]

        # Esto es para poder saber en donde tengo que poner la ultima columna en caso de que cambie el numero de columnas
        size = df.shape[0]
        ##########################
        #  FOR LOOP NEW COLUMN
        ##########################
        # Ahora creamos un forloop para interactuar con todos los rows en estas columnas
        # RECORDAR que no debemos incluir el primer row porque son HEADING y no values
        # Ponemos de limite la variable 'df'
        # Le ponemos el size mas 2 porque me fglatan dos espacio
        # Le decimos que comience desde 2 pero el DF me cuenta solo los spacios con data y los los encabezados
        # Para compensar esto hacemos el truco
        for row in range(2,(size+2)):
            # En cada interaccion pillamos el numero de row para el MES y el año
            # Lo convertimos a STR para poder hacerlo una cadena de texto
            m = str(ws[f'AC{row}'].value)
            y = str(ws[f'AD{row}'].value)

            # Creamos la cadena de texto con la fecha complete
            a_date = "1"+"/"+m+"/"+y

            # En el último parametro especifico el formato de fecha que deseamos tener '%B %Y'
            # B% da el mes con nombre completo
            # m% nos da el numero del mes

            #############################
            #       NEW columna
            #############################
            # En la columna AE ya le hemos asignado nombre
            # Ahora necesitamos asignarle los valores
            ws[f'AE{row}'] = datetime.strptime(a_date, "%d/%m/%Y").strftime('%B %Y')

        #Salvamos el doc como excel
        wb.save(PATH + 'step_1.xlsx')

        df1 = pd.read_excel(PATH +'step_1.xlsx')

        # select  columns to display
        dframe = df1[['Title',
                'Unique_Item_Requests',
                'Total_Item_Requests',
                'Platform',
                'Subject',
                'OrderDescription',
                'OrderNumber',
                'UsedByCustomer',
                'Group',
                'User',
	            'Month',
                'Year',
                'Month-Year']]

        ####################
        #  FIND SIZE NEW DF
        ####################
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        newRows = dframe.shape[0]
        newCols = dframe.shape[1]

        # A esta tabla le quitamos el formato de tabla para incluir la ultima columna que creamos con la fecha
        # Le ponemos el index flase para que no aparezca la coumna index
        dframe.to_excel(PATH + 'step_2.xlsx', index = False)

        return rows, cols, newRows, newCols
#---------- SET UP THE TABLE FORMAT- --------------
    def set_table():

        PATH = 'apps/uploads/tempData/'

        #################
        #  FIND SIZE DF
        #################
        df = pd.read_excel(PATH + 'step_2.xlsx')

        # Aqui buscamos baer el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        #################
        #  UPLOAD BOOK
        #################
        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook(PATH + 'step_2.xlsx')
        # grab the active worksheet
        # This will create the active sheet on this work BOOK
        ws = wb.active

        # Usualmente le ponemos un punto como nombre a la tabla con todos los datos.
        ws.title = "."

        ######################
        #  CREATE TABLE RANGE
        ######################
        # Creamos un objeto de data que guarde los datos de la tabla
        # el parametro 'ref' indica el rango de los datos
        # Necesitamos cojer desde la columna A hasta la M--- porque el el numero de columnas que dejamos
        # A esto le agregamos el numero de rows --SIZE, de esta forma cubrimos todos los datos dentro de la tabla
        tab = Table(displayName='Table1',ref = 'A1:M'+size)

        ##################
        #  STYLE TO TABLE
        ##################
        # Ahora le damos estilo a la tabla
        # Creo una variable que guarde el objeto de estilo
        # Solo quiero que los rows tengan strpes. Las columnas las mantengo normales
        style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn= False,
                                                            showRowStripes=True, showColumnStripes=False)
        # ahora asigno la variable
        tab.tableStyleInfo = style

        # Por ultimo agregamo la tabla al worksheet
        ws.add_table(tab)

        wb.save(PATH + 'step_3.xlsx')

#---------- PIVOT TABLE- --------------
#---------------------------------------
    def test():
        #----- PATH------------
        # Creo un directorio en donde guardar los documentos
        PATH = 'C:\\Users\\Mauro.CespedesAraya\\github\\reportbuilder\\apps\\uploads\\tempData\\webstats\\journals\\'

        #------- CREATE OBJECT---------------
        # Esto es super importante porque aqui creo la instancia de Excel.
        Excel = win32.gencache.EnsureDispatch('Excel.Application',pythoncom.CoInitialize())

        #-------- CREATE BOOK ------------------
        # Creo un Book para guardar los datos de prueba
        # El parametro determina como se cre el nuevo libro
        # String-------------> Busca e nobre y lo utiliza como plantilla para el nuevo libro (ruta de acceso)
        # Contant -----------> El libro continen una solo hajo del tipo espeficado.
                            # XlWBATemplate : xlWBATChart, xlWBATExcel4IntlMacroSheet, xlWBATExcel4MacroSheet o xlWBATWorksheet.
        # No argument -------> Se crea el libro con sheets en blanco
        wb = Excel.Workbooks.Add()

        #------- SELECT SHEET -----------------------
        # El objeto SHEETS devuelve las hojas dentro del workbook
        # Como lo creamos por defecto vamos a especificar el nombre del sheet que deseamos utilizar.
        Sheet1 = wb.Worksheets("Sheet1")
        Sheet1.Name = "."

        #hoja2 = wb.Sheets.Add(After = Sheet1)
        #hoja2.Name = "Journal Stats"

        #---------- CREATE TEST DATA ----------------
        TestData = [['Country','Name','Gender','Sign','Amount'],
                     ['CH','Max' ,'M','Plus',123.4567],
                     ['CH','Max' ,'M','Minus',-23.4567],
                     ['CH','Max' ,'M','Plus',12.2314],
                     ['CH','Max' ,'M','Minus',-2.2314],
                     ['CH','Sam' ,'M','Plus',453.7685],
                     ['CH','Sam' ,'M','Minus',-53.7685],
                     ['CH','Sara','F','Plus',777.666],
                     ['CH','Sara','F','Minus',-77.666],
                     ['DE','Hans','M','Plus',345.088],
                     ['DE','Hans','M','Minus',-45.088],
                     ['DE','Paul','M','Plus',222.455],
                     ['DE','Paul','M','Minus',-22.455]]

        # Aqui recorremos el diccionario para insertar la data en el SHEET1
        for i, TestDataRow in enumerate(TestData):
            for j, TestDataItem in enumerate(TestDataRow):
                # Utilizamos el metodo Cells
                # Indicamos como argumento el indice de ROW y COLUMN
                # Si no indicamos parametro representa toda la hoja de excel.
                        #Sheet1.Cells(i+2,j+4).Value = TestDataItem
                # De est forma pone la table desde el primer ROW y la primera COL
                Sheet1.Cells(i+1,j+1).Value = TestDataItem

        #-------- SELECT TABLE ------------------
                        #cl1 = Sheet1.Cells(2,4)
                        #cl2 = Sheet1.Cells(2+len(TestData)-1,4+len(TestData[0])-1)

        # Aún no entiendo esto al 100% pero selecciono el mismo range que el FOR forloop
        # Esta es la informacion que le pasamos al pivot table
        # Es como definir lo rangos que deseamos que tome
        cl1 = Sheet1.Cells(1,1)
        cl2 = Sheet1.Cells(1+len(TestData)-1,1+len(TestData[0])-1)

        #############################################################################################################
        #                           FIRST PIVOT TABLE
        #############################################################################################################
        #-------- CREATE PIVOT TABLE
        # Aqui le pasamos los rangos a una variable
        PivotSourceRange = Sheet1.Range(cl1,cl2)
        # Esta variable queda seleccionada.
        PivotSourceRange.Select()

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        hoja2 = wb.Sheets.Add(After = Sheet1)
        hoja2.Name = "Journal Stats"

        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el 2, porque queremos asignar la hoja dos que acabamos de crear.
        Sheet2 = wb.Worksheets(2)

        # Una vez seleccionada la data le indicamos en donde deseamos colocar el pivot_table
        # Según conversacones con Jeff comenzamos en el row 7 de la primera columna.
        cl3=Sheet2.Cells(7,1)

        # Ahora insertamos esos datos dentro del rango seleccionado.
        PivotTargetRange=  Sheet2.Range(cl3,cl3)
        PivotTableName="ReportPivotTable"

        # Esto aún no lo entiendo bien pero es la forma de inicializar la pivot table
        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)
        #-----------------#
        #        ROW
        #-----------------#
        PivotTable.PivotFields('Name').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Name').Position = 1
        #-----------------#
        #        FILTER
        #-----------------#
        PivotTable.PivotFields('Gender').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Gender').Position = 1
        # Esto es en caso que querramos tener algo seleccionado por default
        #PivotTable.PivotFields('Gender').CurrentPage="M"
        #-----------------#
        #        COLUMNS
        #-----------------#
        #PivotTable.PivotFields('Country').Orientation = win32c.xlColumnField
        #PivotTable.PivotFields('Country').Position = 1
        #
        #PivotTable.PivotFields('Country').Subtotals = [False, False, False, False, False, False, False, False, False, False, False, False]
        #PivotTable.PivotFields('Sign').Orientation = win32c.xlColumnField
        #PivotTable.PivotFields('Sign').Position = 2
        #-----------------#
        #        VALUES
        #-----------------#
        DataField = PivotTable.AddDataField(PivotTable.PivotFields('Amount'))
        DataField.NumberFormat="#"##0.00'

        wb.SaveAs(PATH + 'trickTest.xlsx')
        # Esto es en caso que deseemos hacer el Excel visible mientras lo trabajamos
        # En este caso prefiero dejarlo como que no se muestre.
        Excel.Visible = False

        #-------- SAVE WORKBOOK
        wb.SaveAs(PATH + 'trickTest.xlsx')

        #-------- QUIT EXCEL
        Excel.Application.Quit()

    def test1():
        #----- PATH------------
        # Creo un directorio en donde guardar los documentos
        PATH = 'C:\\Users\\Mauro.CespedesAraya\\github\\reportbuilder\\apps\\uploads\\tempData\\webstats\\journals\\'
        #------- CREATE OBJECT---------------
        # Esto es super importante porque aqui creo la instancia de Excel.
        Excel = win32.gencache.EnsureDispatch('Excel.Application',pythoncom.CoInitialize())

        #-------- OPEN BOOK ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = Excel.Workbooks.Open('C:\\Users\\Mauro.CespedesAraya\\github\\reportbuilder\\apps\\uploads\\tempData\\webstats\\journals\\trickTest.xlsx')

        #------- SELECT SHEET -----------------------
        # El objeto SHEETS devuelve las hojas dentro del workbook
        # Como lo creamos por defecto vamos a especificar el nombre del sheet que deseamos utilizar.
        Sheet1 = wb.Worksheets(".")
        Sheet1 = wb.ActiveSheet

        #-------- SELECT TABLE ------------------
                        #cl1 = Sheet1.Cells(2,4)
                        #cl2 = Sheet1.Cells(2+len(TestData)-1,4+len(TestData[0])-1)

        # Aún no entiendo esto al 100% pero selecciono el mismo range que el FOR forloop
        # Esta es la informacion que le pasamos al pivot table
        # Es como definir lo rangos que deseamos que tome
        cl1 = Sheet1.Cells(1,5)
        cl2 = Sheet1.Cells(1,5)
        #-------- CREATE PIVOT TABLE
        # Aqui le pasamos los rangos a una variable
        PivotSourceRange = Sheet1.Range(cl1,cl2)
        # Esta variable queda seleccionada.
        PivotSourceRange.Select()

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        hoja3 = wb.Sheets.Add(After = Sheet1)
        hoja3.Name = "test"

        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el 2, porque queremos asignar la hoja dos que acabamos de crear.
        Sheet2 = wb.Worksheets(3)

        # Una vez seleccionada la data le indicamos en donde deseamos colocar el pivot_table
        # Según conversacones con Jeff comenzamos en el row 7 de la primera columna.
        cl3=Sheet2.Cells(1,7)

        # Ahora insertamos esos datos dentro del rango seleccionado.
        PivotTargetRange=  Sheet2.Range(cl3,cl3)
        PivotTableName="ReportPivotTable"

        # Esto aún no lo entiendo bien pero es la forma de inicializar la pivot table
        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)
        #-----------------#
        #        ROW
        #-----------------#
        PivotTable.PivotFields('Name').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Name').Position = 1
        #-----------------#
        #        FILTER
        #-----------------#
        PivotTable.PivotFields('Gender').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Gender').Position = 1
        # Esto es en caso que querramos tener algo seleccionado por default
        #PivotTable.PivotFields('Gender').CurrentPage="M"
        #-----------------#
        #        COLUMNS
        #-----------------#
        #PivotTable.PivotFields('Country').Orientation = win32c.xlColumnField
        #PivotTable.PivotFields('Country').Position = 1
        #
        #PivotTable.PivotFields('Country').Subtotals = [False, False, False, False, False, False, False, False, False, False, False, False]
        #PivotTable.PivotFields('Sign').Orientation = win32c.xlColumnField
        #PivotTable.PivotFields('Sign').Position = 2
        #-----------------#
        #        VALUES
        #-----------------#
        DataField = PivotTable.AddDataField(PivotTable.PivotFields('Amount'))
        DataField.NumberFormat="#"##0.00'

        # Esto es en caso que deseemos hacer el Excel visible mientras lo trabajamos
        # En este caso prefiero dejarlo como que no se muestre.
        Excel.Visible = False

        #-------- SAVE WORKBOOK
        wb.SaveAs(PATH + '2trickTest.xlsx')

        #-------- QUIT EXCEL
        Excel.Application.Quit()


    def set_graphic():

        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook('docs/temp/final_table.xlsx')
        ws = wb.active

        df = pd.read_excel('docs/temp/final_table.xlsx')

        size = df.shape[0]
        print(size)
        #-------------------------------------
        # Creamos otro tab pra guardar el grafico
        # With teh work book object we call the create fucntion
        ws_1 = wb.create_sheet("Statistics")

        #-------------------------------------
        # CREAMOS grafico
        #-------------------------------------

        # Ahora que tenemos informacion tenemos que decidir como se presenta en la grafica
        chart = BarChart()

        # Especificamos desde donde, hasta donde queremos la data
        # Tenemos que encontrar el tamaño de la tabla
        labels = Reference(ws,min_col=2, min_row=1, max_row=size)

        data = Reference(ws,min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)


        chart.add_data(data,titles_from_data=True)
        chart.set_categories(labels)
        chart.title = 'Ice crema Flavor'

        # Añadimos el gráfico a la pestaña correcta
        # Decimos en que celda deseo que comience.
        ws_1.add_chart(chart,'Q5')

        # Cunado tenemos todo lo que queremos guardamos el archivo
        wb.save('docs/temp/graphs.xlsx')

        return wb
