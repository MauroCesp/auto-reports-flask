try:
    # Voy a utilizar dos formas de importar paquetes
    # Primero importo solo lo que necesito directamente de cada sub-package
    # En este caso importo los modulos de cada uno para utilizar sus funciones
    # Estos son los imports para crear pivot tables
    from pywintypes import com_error
    from pathlib import Path
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from datetime import datetime, date
    from openpyxl import Workbook
    # Esto libreria nos ayuda a crear pivot tables
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    # Esto es para las fechas

    #------------------------------------------------------------------------------
    from openpyxl.styles import NamedStyle
    from openpyxl.styles.numbers import builtin_format_code, builtin_format_id
    from openpyxl.styles.numbers import is_builtin, is_date_format, is_datetime
    from openpyxl.styles import Color, PatternFill, Font, Border

    #----------  IMPORT APP -----------------------------------
    # Es super necesario llamar al APP para poder utilizar las libreria propia
    # Como cunado necesitamos llamar al directorio raiz de la applicacion a la hora de importar ficheros.
    from run import app
    #------------- Tablas e imagenes -------------------------
    # Primero importo algunas funciones para tablas e imagenes
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import PieChart, Reference, Series,PieChart3D,LineChart, BarChart
    # Ahora utilizo directamente el import para llamar a un modulo que no pertenece a ningún sub-paquete
    # El modulo se encuetra en el mismo nivel que los sub-paquetes
    #------------ PIvot tables -------------------------------
    # Estos son los imports para crear pivot tables
    import win32com.client as win32
    win32c = win32.constants
    import sys
    import numpy as np  # only used for synthetic data
    import random  # only used for synthetic data
    import pythoncom
    import xlwings as xw
    import time
    import pandas as pd
    from PIL import Image
    import os
    import xlsxwriter
    import calendar
# Utilizo la excepción específica para módulos
except ModuleNotFoundError as err:
    print('Opssss... Looks like there is an error importing the package', err)
################################################################################

class Book():
    #------------- CONSTRUCTOR ------------------
    # Inicializo el constructor
    # Cada vez que cree un objeto de tipo report se inicializan estos atributos
    # Se repiten en todos los metodos aunque cambia el valores# El valor lo cambio directo desde el archivo de rutas
    def __init__(self, TEMP_SAVE_PATH, FINAL_SAVE_PATH, SOURCE_FILE, type_graphic, use_grid, session_select,TII_select, TIR_select, UIR_select , Monthly_Trend, Yearly_Comparison, Monthly_Trend1, SR_select, image, UII_select):
        # Definimos el tipo de grafico que queremos
        self.type_graphic = type_graphic
        # Aqui determinamos si el grafico utiliza grids o no
        self.use_grid = use_grid
        self.pic = image

        # Creamos atributos de instancia
        self.FINAL_SAVE_PATH = FINAL_SAVE_PATH
        self.TEMP_SAVE_PATH = TEMP_SAVE_PATH

        self.session_select = session_select
        self.TII_select = TII_select
        self.TIR_select = TIR_select
        self.UIR_select = UIR_select
        self.UII_select = UII_select
        self.Monthly_Trend = Monthly_Trend
        self.Yearly_Comparison = Yearly_Comparison
        self.Monthly_Trend1 = Monthly_Trend1


        # Cunado creamos las aplicacion de dentro de Excel necesitamos utilizar absolute path
        #-------------- Aabspath ----------------------#
        # Para que el codigo funcione en cualquier maquina debemos utilizar la funcion de ABSPA
        # Esta es la ruta final pero de los archivos temporales , los tenemos que moer a un archivo permanente.
        self.AB_FINAL_SAVE_PATH = os.path.abspath('apps/uploads/tempData/webstats/books/')
        self.AB_FINAL_SAVE_PATH_PERMANENT = os.path.abspath('apps/uploads/finalData/webstats/books/')

        self.AB_TEMP_SAVE_PATH = os.path.abspath('apps/uploads/tempData/')
        # Aqui es don de se van a guardar las pics
        self.AB_PIC_PATH = os.path.abspath('apps/static/assets/img/company-logos/')

        #----------- Inicializo objeto de Windows
        self.Excel = win32.gencache.EnsureDispatch('Excel.Application',pythoncom.CoInitialize())

        #---------- CLEAN Data
        # Vamos a quitar todos los campos que no tienen SESSION & TOTAL ITEM INVESTIGATIONS
        df = pd.read_excel(self.TEMP_SAVE_PATH + SOURCE_FILE)
        # Si ambas columnas no tienen datos borro la linea
        df1 = df[(df['Sessions'] != 0) & (df['Total_Item_Investigations'] != 0)]
        df1.to_excel(TEMP_SAVE_PATH + 'cleaned_Data.xlsx', index=None)
        self.SOURCE_FILE = TEMP_SAVE_PATH + 'cleaned_Data.xlsx'

        # ---------   FIND COMPANY NAME ----------------
        # Saco el nombre de la compañia del primer archivo para no liarme mucho
        df1 = pd.read_excel(self.TEMP_SAVE_PATH + 'cleaned_Data.xlsx', engine='openpyxl')

        #Busco el nombre del cliente

        # lAS IMAGENENS NO ACEPTAN / en el nombre. ENtonces si recibimos una imagen nunca la econtraremos porque elimina el /
        # Lo que hacemos es quitr del nombre de la institucion cualquier / tambien.
        cnameR = str(df1.loc[0].at['UsedByCustomer'] )
        # Quito lo espacios porque sino me da problemas al final a l tratar de leer el archivo
        cname = cnameR.replace("/", "")

        cname1 = cname.replace(' ', '_')
        self.cName = cname1

        # Esto es para buscar las imagenes en linea sin los '_'
        self.cName1 = cname

        # ---------   FIND YEARS of report----------------
        self.cYear = df1.Year.unique()

        #sort values smallest to largest
        self.cYear.sort()
        self.startYear = str(self.cYear[0])
        self.closeYear = str(self.cYear[-1])
        #--------------FIND Months of report-------------------------

        # Aqui creo dos datasets nuevos para dividir los años de inicio y final
        # Ahora busco el mes minimo en cada uno de ellos

        initY = df1.loc[(df1['Year']==self.cYear[0])]


        endY = df1.loc[(df1['Year']==self.cYear[-1])]

        # Creo un datafrme solo con la columna de mes para poder filtrarlo
        initY_1 = initY['Month']

        endY_1 = endY['Month']


        # COmo está en números busco el numero menor que seria igual al primer mes
        initY_2  = initY_1.min()
        endY_2  = endY_1.max()


        # https://pynative.com/python-get-month-name-from-number/#:~:text=Use%20the%20calendar.,year%20in%20the%20current%20locale.
        self.initMonth  = calendar.month_name[initY_2]
        self.endMonth = calendar.month_name[endY_2]


        #---------------------------------------------------------
        #  ESTO ES PARA SABER SI SON AÑOS COMPLETOS O NO
        #---------------------------------------------------------

        # Esto es lo que voy a agarrar desde los metodos de average y TREND para saber si
        # un año es completo o no. SOlo mostrara si los años son completos, porque sino se ve muy mal.
        total = []
        for i in self.cYear:
            x = df.loc[(df['Year']==i)]

            total.append(x.Month.nunique())

        f = []
        for i in total:
            if i == 12:
                f.append(True)
            else:
                f.append(False)

        # Convierto las dos listas en un diccionario
        self.complete_year = dict(zip(self.cYear, f))


    def excelQuit(self):

        self.Excel.Application.Quit()
#---------- CREATE NEW COLUMN -MONTH-YEAR- --------------
    def new_column(self):
        ##################
        #  OPEN FILE
        ##################
        # Ahora vamos a buscar el archivo y cargarlo
        wb = load_workbook(self.SOURCE_FILE)
        ws = wb.active

        # Ahora creamos la cell que queremos agregar
        # Hay que tener en cuanta el tamaño del sheet para poder ubicarla bien en la columna correcta
        # Creo una variable donde guardo la cell

        newCol = ws['AE1']
        newCol.value = 'Month-Year'

        #################
        #  FIND SIZE DF
        #################
        # Podemos averiguar el tamaño de la coumna facilment con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(self.SOURCE_FILE, engine='openpyxl')

        rows = df.shape[0]
        cols = df.shape[1]
        # Esto es para poder saber en donde tengo que poner la ultima columna en caso de que cambie el numero de columnas
        size = df.shape[0]
        ##########################
        #  FOR LOOP NEW COLUMN
        ##########################
        # Ahora creamos un forloop para interactuar con todos los rows en estas columnas
        # RECORDAR que no debemos incluir el primer row porque son HEADING y no values
        # Ponemos de limite la variable 'df'
        # Le ponemos el size mas 2 porque me fglatan dos espacio
        # Le decimos que comience desde 2 pero el DF me cuenta solo los spacios con data y los los encabezados
        # Para compensar esto hacemos el truco

        for row in range(2,(size+2)):
            # En cada interaccion pillamos el numero de row para el MES y el año
            # Lo convertimos a STR para poder hacerlo una cadena de texto
            m = str(ws[f'AC{row}'].value)
            y = str(ws[f'AD{row}'].value)

            # Creamos la cadena de texto con la fecha complete
            a_date = "1"+"/"+m+"/"+y

            # En el último parametro especifico el formato de fecha que deseamos tener '%B %Y'
            # B% da el mes con nombre completo
            # b% nos da la abreviatura del mes
            # m% nos da el numero del mes

            #############################
            #       NEW columna
            #############################
            # En la columna AE ya le hemos asignado nombre
            # Ahora necesitamos asignarle los valores
            ws[f'AE{row}'] = datetime.strptime(a_date, "%d/%m/%Y").strftime('%b-%Y')

        #Salvamos el doc como excel
        wb.save(self.TEMP_SAVE_PATH + 'step_1.xlsx')

        df1 = pd.read_excel(self.TEMP_SAVE_PATH +'step_1.xlsx')

        # select  columns to display
        dframe = df1[['Title',
                'Unique_Item_Requests',
                'Total_Item_Requests',
                'Total_Item_Investigations',
                'Unique_Item_Investigations',
                'OA_Gold',
                'Sessions',
                'Platform',
                'Subject',
                'OrderDescription',
                'OrderNumber',
                'UsedByCustomer',
                'Group',
                'User',
	            'Month',
                'Year',
                'Month-Year']]

        ####################
        #  FIND SIZE NEW DF
        ####################
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        newRows = dframe.shape[0]
        newCols = dframe.shape[1]

        #dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime('%b-%Y')
        #dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime("%Y-%m-%d")

        dframe['Month-Year'] = pd.to_datetime(dframe['Month-Year']).dt.strftime("%b-%Y")

        # A esta tabla le quitamos el formato de tabla para incluir la ultima columna que creamos con la fecha
        # Le ponemos el index flase para que no aparezca la coumna index
        dframe.to_excel(self.TEMP_SAVE_PATH + 'step_2.xlsx', index = False)

        self.Excel.Application.Quit()
        # Enviamos el nombre de la empresa para buscar la imagen.
        return self.cName

#---------- SET UP THE TABLE FORMAT- --------------
    def set_table(self):

        #################
        #  FIND SIZE DF
        #################
        df = pd.read_excel(self.TEMP_SAVE_PATH + 'step_2.xlsx')

        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        #################
        #  UPLOAD BOOK
        #################
        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook(self.TEMP_SAVE_PATH + 'step_2.xlsx')
        # grab the active worksheet
        # This will create the active sheet on this work BOOK
        ws = wb.active

        # Usualmente le ponemos un punto como nombre a la tabla con todos los datos.
        ws.title = "."

        ######################
        #  CREATE TABLE RANGE
        ######################
        # Creamos un objeto de data que guarde los datos de la tabla
        # el parametro 'ref' indica el rango de los datos
        # Necesitamos cojer desde la columna A hasta la M--- porque es el numero de columnas que dejamos
        # A esto le agregamos el numero de rows --SIZE, de esta forma cubrimos todos los datos dentro de la tabla
        tab = Table(displayName='Table1',ref = 'A1:Q'+size)

        ##################
        #  STYLE TO TABLE
        ##################
        # Ahora le damos estilo a la tabla
        # Creo una variable que guarde el objeto de estilo
        # Solo quiero que los rows tengan strpes. Las columnas las mantengo normales
        style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn= False,
                                                            showRowStripes=True, showColumnStripes=False)
        # ahora asigno la variable
        tab.tableStyleInfo = style

        # Por ultimo agregamos la tabla al worksheet
        ws.add_table(tab)

        wb.save(self.TEMP_SAVE_PATH + 'step_3.xlsx')
        wb.close()

        return wb
        self.Excel.Application.Quit()

#---------- Montly Trend- --------------
    def pivot_montly_trend(self):

        source_file = self.AB_TEMP_SAVE_PATH + '\\step_3.xlsx'

        self.Excel.Visible = False
        #-------- OPEN BOOK ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = self.Excel.Workbooks.Open(source_file)

        # Esta es la forma de seleccionar el SHEET dentro de work book.
        ws = wb.Worksheets('.')

        #------  FIND SIZE DF --------------------
        # Podemos averiguar el tamaño de la columna facilmente con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(source_file, engine='openpyxl')
        rows = df.shape[0]
        cols = df.shape[1]
        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        #------  SELECT RANGE  --------------------
        # Pongo M porque se que tengo 13 columnas que son la que deje despues de filtrar la informacion
        # Como el numero de datos puede variar necesito hacerlo dinamico.
        PivotSourceRange = ws.Range('A1:Q'+size)

        #############################################################################################################
        #                           FIRST PIVOT TABLE
        #############################################################################################################
        #-------- CREATE PIVOT TABLE
        # Aqui le pasamos los rangos a una variable
        #PivotSourceRange = wb.Range('A1:M'+size)

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        hoja3 = wb.Sheets.Add()
        hoja3.Name = "Monthly Trend"


        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el 1, porque queremos asignar la hoja dos que acabamos de crear.
        # La hoja original se va al final
        Sheet3 = wb.Worksheets(1)

        # --------- HIDE Grind --------------

        if self.use_grid == 'True':
            self.Excel.ActiveWindow.DisplayGridlines = True
        else:
            self.Excel.ActiveWindow.DisplayGridlines = False
        # --------- HIDE Grind --------------

        ############ CHECK POINT------> Hasta aqui todo bien

        # Una vez seleccionada la data le indicamos en donde deseamos colocar el pivot_table
        # Según conversacones con Jeff comenzamos en el row 7 de la primera columna.
        cl3=Sheet3.Cells(7,1)

        # Ahora insertamos esos datos dentro del rango seleccionado.
        PivotTargetRange=  Sheet3.Range(cl3,cl3)
        PivotTableName="ReportPivotTable"

        # Esto aún no lo entiendo bien pero es la forma de inicializar la pivot table
        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)

        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)
        #-----------------#
        #        ROW
        #-----------------#
        PivotTable.PivotFields('Year').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Year').Position = 1
        # Esto es para eliminar los subtotals
        #PivotTable.PivotFields('Year').Subtotals = tuple(False for _ in range(12))

        PivotTable.PivotFields('Month').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Month').Position = 2
        # Esto es para eliminar los subtotals
        #PivotTable.PivotFields('Month').Subtotals = tuple(False for _ in range(12))

        # Esto es para que no me aparezca el contenido
        #PivotTable.PivotFields('Month').DataRange.NumberFormat = ";;;"

        PivotTable.PivotFields('Month-Year').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Month-Year').Position = 3

        #-----------------#
        #        FILTER
        #-----------------#
        PivotTable.PivotFields('Title').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Title').Position = 1
        # Esto es en caso que querramos tener algo seleccionado por default
        #PivotTable.PivotFields('Gender').CurrentPage="M"

        #-----------------#
        #        VALUES (Columnas)
        #-----------------#

        # Auiq lo que hago es crear un array para guardar todas las opciones que han sido seleccinadas
        # Evaluo la condicion uno por uno y si existe lo voy metiendo en el array para trabajar es STYLE
        selection_list = []

        if  self.session_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Sessions'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Sessions')
        else:
            pass

        if self.TII_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Investigations'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Investigations')
        else:
            pass

        if self.TIR_select == 'Total_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Requests'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Requests')
        else:
            pass

        if self.UIR_select == 'Unique_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Requests'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Requests')
        else:
            pass

        if self.UII_select == 'Unique_Item_Investigations':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Investigations'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Investigations')
        else:
            pass
        #-----------------#
        #       FORMAT
        #-----------------#
        # Esto es para que aparezca striped
        PivotTable.ShowTableStyleRowStripes = True
        # Genera una pequeña linea para los Headers
        PivotTable.ShowTableStyleColumnHeaders = True

        #------------------------#
        #        Title of Pivot
        #------------------------#
        Book.customize_title(self,Sheet3)

        #--------------------------#
        #       CHANGE FIELD NAMES
        #-------------------------#
        arg1 = 'Year and Month '
        arg4 = 'Book'

        # Utilizo el metodo style para cambiar el  Estilo y nombre de los campos
        Book.style(self, Sheet3, arg1, selection_list, arg4, size)

        #--------------------------#
        #       CRETE CHART
        #-------------------------#
        #Book.create_chart(self, Sheet3)

        #--------------------------#
        #       HIDE ROWS
        #-------------------------#

        lastrow = Sheet3.UsedRange.Rows.Count
        lastrow = lastrow+8

        dict =['2018.0','2019.0','2020.0','2021.0','2022.0','1.0','2.0','3.0','4.0','5.0','6.0','7.0','8.0','9.0','10.0','11.0','12.0']

        for i in range(8, lastrow):
            #print('Este es el row:',Sheet3.Cells(i,1))
            a  = Sheet3.Cells(i,1)
            x = Book.__repr__(self,a)

            for j in dict:
                if x == j:
                    Sheet3.Rows(i).EntireRow.Hidden = True
                else:
                    continue

#---------------------------------------------------------------------------------------

        #-------- SAVE WORKBOOK
        wb.SaveAs(self.AB_FINAL_SAVE_PATH + '\\1.xlsx')

        time.sleep(2)
        #-------- QUIT EXCEL
        self.Excel.Application.Quit()

#---------- READ as STRING- --------------
    def __repr__(self, x):
        #A class can control what this function returns for its instances by defining a __repr__() method.
        # Aqui lo que digo es que quiero que larepresentacion de X sea de forma string
        return  str(x)

#---------- PIVOT TABLE- --------------
    def pivot_stats(self):

        source_file = self.AB_TEMP_SAVE_PATH + '\\step_3.xlsx'

        self.Excel.Visible = False
        #-------- OPEN Book ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = self.Excel.Workbooks.Open(source_file)

        # Esta es la forma de seleccionar el SHEET dentro de work Book.
        ws = wb.Worksheets('.')

        #------  FIND SIZE DF --------------------
        # Podemos averiguar el tamaño de la columna facilmente con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(source_file, engine='openpyxl')
        rows = df.shape[0]
        cols = df.shape[1]
        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        #------  SELECT RANGE  --------------------
        # Pongo O porque se que tengo 15 columnas que son la que deje despues de filtrar la informacion
        # Como el numero de datos puede variar necesito hacerlo dinamico.
        PivotSourceRange = ws.Range('A1:Q'+size)

        #############################################################################################################
        #                           FIRST PIVOT TABLE
        #############################################################################################################
        #-------- CREATE PIVOT TABLE
        # Aqui le pasamos los rangos a una variable
        #PivotSourceRange = wb.Range('A1:M'+size)

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        hoja3 = wb.Sheets.Add()
        hoja3.Name = "Book Stats"

        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el , porque queremos asignar la hoja dos que acabamos de crear.
        Sheet3 = wb.Worksheets(1)

        # --------- HIDE Grind --------------

        if self.use_grid == 'True':
            self.Excel.ActiveWindow.DisplayGridlines = True
        else:
            self.Excel.ActiveWindow.DisplayGridlines = False
        # --------- HIDE Grind --------------

        ############ CHECK POINT------> Hasta aqui todo bien

        # Una vez seleccionada la data le indicamos en donde deseamos colocar el pivot_table
        # Según conversacones con Jeff comenzamos en el row 7 de la primera columna.
        cl3=Sheet3.Cells(7,1)
        #------------------------#
        #        Title of Pivot
        #------------------------#
        Book.customize_title(self,Sheet3)

        # Ahora insertamos esos datos dentro del rango seleccionado.
        PivotTargetRange=  Sheet3.Range(cl3,cl3)
        PivotTableName="ReportPivotTable"

        # Esto aún no lo entiendo bien pero es la forma de inicializar la pivot table
        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)

        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

        #-----------------#
        #        ROW
        #-----------------#
        PivotTable.PivotFields('Title').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Title').Position = 1
        #-----------------#
        #        FILTER
        #-----------------#
        PivotTable.PivotFields('Year').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Year').Position = 1
        # Esto es en caso que querramos tener algo seleccionado por default
        #PivotTable.PivotFields('Gender').CurrentPage="M"
        #-----------------#
        #        VALUES (Columnas)
        #-----------------#

        # Auiq lo que hago es crear un array para guardar todas las opciones que han sido seleccinadas
        # Evaluo la condicion uno por uno y si existe lo voy metiendo en el array para trabajar es STYLE
        selection_list = []

        if  self.session_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Sessions'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Sessions')
        else:
            pass

        if self.TII_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Investigations'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Investigations')
        else:
            pass

        if self.TIR_select == 'Total_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Requests'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Requests')
        else:
            pass

        if self.UIR_select == 'Unique_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Requests'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Requests')
        else:
            pass

        if self.UII_select == 'Unique_Item_Investigations':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Investigations'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Investigations')
        else:
            pass
        #-----------------#
        #       FORMAT
        #-----------------#
        # Esto es para que aparezca striped
        PivotTable.ShowTableStyleRowStripes = True

        #--------------------------#
        #       CHANGE FIELD NAMES
        #-------------------------#
        arg1 = 'Book'
        arg4 = 'Year'

        # Utilizo el metodo style para cambiar el  Estilo y nombre de los campos
        Book.style(self, Sheet3, arg1, selection_list, arg4, size)
        #-------- SAVE WORKBOOK
        #Sheet3.Columns.AutoFit()

        #------------------------#
        #     SORT
        #------------------------#
        #Book.sort(self,Sheet3)

        #wb.SaveAs(self.AB_FINAL_SAVE_PATH + self.cName + ' ' + 'Book_Stats' + ' '+ self.startYear +'-'+ self.closeYear +'.xlsx')
        wb.SaveAs(self.AB_FINAL_SAVE_PATH + '\\2.xlsx')

        time.sleep(2)
        #-------- QUIT EXCEL
        self.Excel.Application.Quit()

    def RGB(r,g,b):

        bgr = (r, g, b)
        strValue = '%02x%02x%02x' % bgr
        # print(strValue)
        iValue = int(strValue, 16)
        return iValue

    def style(self, Sheet3, arg1, selection_list, arg4, size):
        #--------------------------#
        #       CHANGE FIELD NAMES
        #-------------------------#
        Sheet3.Cells(7,1).Value = arg1
        Sheet3.Cells(7,1).Font.Size = 13
        Sheet3.Cells(7,1).Font.Bold = True
        Sheet3.Cells(7,1).Font.Color = Book.RGB(0,0,0)
        Sheet3.Cells(7,1).HorizontalAlignment = win32c.xlCenter

        # Este counter es para colocar cada columna detras de la otr de acuerdo a cuantas hayan
        counter = 2

        # Es un truco para que me guarda los nombre de las variables bien
        test = ' '

        # Interactuo sobre cada elemento y voy creando columnas
        for i in selection_list:

            x = i.replace ('_', ' ')
            # LE meto TEST a la i para que sea diferente en cada vuelta sino me da error.
            Sheet3.Cells(7,counter).Value = x+test
            Sheet3.Cells(7,counter).Font.Size = 13
            Sheet3.Cells(7,counter).Font.Bold = True
            Sheet3.Cells(7,counter).Font.Color = Book.RGB(0,0,0)
            Sheet3.Cells(7,counter).HorizontalAlignment = win32c.xlCenter

            # Sumo al counter para ir cambiando de posicion
            counter += 1
            # Sumo al test para que sea diferente en cada vuelta.
            test = test+''

        Sheet3.Cells(5,1).Value = arg4
        Sheet3.Cells(5,1).Font.Bold = True
        Sheet3.Cells(5,1).Font.Size = 13
        Sheet3.Cells(5,1).Font.Color = Book.RGB(0,0,0)
        Sheet3.Cells(5,1).HorizontalAlignment = win32c.xlCenter

        # ------------------------------
        # -------- LAST (COL - ROW) ----
        # ------------------------------
        # Esto nos sirve para averiguar cual es el ultimo campo utilizado en la tabla.
        lastrow = str(Sheet3.UsedRange.Rows.Count)
        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet3.UsedRange.Columns.Count
        # Asi convierto el numero de columna en character
        # Le sumo 1 porque la primera columna no es parte de los resultados .
        # SI la dejo asi me cuenta una columna menos
        letter = chr(ord('@') + int(lastcol)+1)


        # Adding borders
        # Escojo el range que deseo ttrabajar
        # Agrego 2 espacios más porque parece que no me cubre todo
        # Borders ---> 1 significa borde solo de un lado, 2 significa borde doble(ambos lados de la columna)
        # Weight ---> 1 es linea punteada, 2 es linea solida
        ########Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(size)+2),3)).Borders(2).Weight = 1

        # Dee sta forma hacemos que los datos dentro del pivot esten centrados
        #Comienzo desde la segunda columna para mantener los titles a la derecha por defecto.
        Sheet3.Range(Sheet3.Cells(7,2),Sheet3.Cells((1+int(lastrow)),(1+lastcol))).HorizontalAlignment = win32c.xlCenter


        # ----------------------------------------------------------------
        #           EDGES
        # ----------------------------------------------------------------
        # Primero el filtro de arriba le ponermos grid
        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)

        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlInsideHorizontal).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(5,1),Sheet3.Cells(5,2)).Borders(win32c.xlInsideVertical).Color = Book.RGB(0, 0, 0)

        #---- aHORA A TODO EL pIVOT LE DAMOS grid
        #https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)

        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlInsideHorizontal).Color = Book.RGB(0, 0, 0)
        Sheet3.Range(Sheet3.Cells(7,1),Sheet3.Cells((int(lastrow)),(lastcol))).Borders(win32c.xlInsideVertical).Color = Book.RGB(0, 0, 0)


        #------  SELECT RANGE  --------------------
        # Pongo M porque se que tengo 13 columnas que son la que deje despues de filtrar la informacion
        # Como el numero de datos puede variar necesito hacerlo dinamico.
        PivotSourceRange = Sheet3.Range('A8:'+letter+lastrow)

    def customize_title(self, Sheet3):
        #-------- Add title to table

        cname1 = self.cName.replace('_', ' ')

        Sheet3.Cells(3,1).Value = cname1 + ' ' + 'Book Stats' + ' '+ '  ('+ self.initMonth +' '+self.startYear +' - '+' '+ self.endMonth + ' '+self.closeYear+') '

        # ------------------------------
        # -------- LAST (COL - ROW) ----
        # ------------------------------
        # Esto nos sirve para averiguar cual es el ultimo campo utilizado en la tabla.
        lastrow = str(Sheet3.UsedRange.Rows.Count)
        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet3.UsedRange.Columns.Count
        # Asi convierto el numero de columna en character
        letter = chr(ord('@') + int(lastcol)-1)


        # ------- Merge cells
        # Escojo el rango de cells
        # El primer numero es rows# Segundo numero es column
        #myRange = Sheet3.Range(Sheet3.Cells(3,1),Sheet3.Cells(3,lastcol))
        #myRange.MergeCells = True
        # Las options de align son : Center, Right,Left
        #myRange.HorizontalAlignment = win32c.xlLeft
        # -------How to add colors
        #https://itecnote.com/tecnote/python-pywin32-excel-formatting/
        # myRange.Interior.ColorIndex = 6
        #Shet3.Range("A1", "D1").Interior.ColorIndex = 6

        #---- Customize title
        Sheet3.Range('A3','A'+str(lastcol)).Font.Size = 15
        Sheet3.Range('A3','A'+str(lastcol)).Font.Bold = True
        #https://www.w3schools.com/colors/colors_picker.asp
        Sheet3.Range('A3','A'+str(lastcol)).Font.Color = Book.RGB(0, 0, 0)

    def create_chart(self, Sheet3):

        # ------------------------------
        # -------- LAST (COL - ROW) ----
        # ------------------------------
        # Esto nos sirve para averiguar cual es el ultimo campo utilizado en la tabla.
        lastrow = Sheet3.UsedRange.Rows.Count
        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet3.UsedRange.Columns.Count
        # Asi convierto el numero de columna en character
        letter = chr(ord('@') + int(lastcol)-1)

        Sheet3.Range('A7:'+letter+str(2+lastrow)).Select()

        # --------- HIDE Grind --------------

        if self.use_grid == 'True':
            self.Excel.ActiveWindow.DisplayGridlines = True
        else:
            self.Excel.ActiveWindow.DisplayGridlines = False

        # --------------------------------------
        #  SELECT TYPE OF GRAPH
        # --------------------------------------
        # Al opcion ADDCHART ofrece muchas posibilidades de customizacion
        # 1- Tipo de grafico-----> solo tenemos que decidir el tipo de grafico que deseamos utilizar
            ##https://docs.microsoft.com/es-es/office/vba/api/excel.xlcharttype
                    #Excel.ActiveChart.ChartType = win32c.xlLine
        # 2- LEFT------> Escojemos cuanto queremos que se aleje el grafico del borde izquierdo (Se mide en puntos)
        # 3- TOP  ----> Escojemos cunato se aleja el grafico del borde superior (Se mide en puntos)
        # 4- WIDTH ---> (Se mide en puntos)
        # 5- HEIGHT ---> (Se mide en puntos)

        ###############################################################
        #------------ IMAGE ----------------------
        ###############################################################
        # https://learn.microsoft.com/es-es/office/vba/api/excel.shapes.addpicture
        # 1- Path-----> definimos donde está el archivo
                        # Eventualmente será necesario hacer una base de datos referencial de imagenes.
        # TRUE --- vincular la imagen al archivo desde el que se creó.

        # True --------> guardar la imagen vinculada con el documento en el que se inserta.
        # 2- LEFT------> Escojemos cuanto queremos que se aleje el grafico del borde izquierdo (Se mide en puntos)
        # 3- TOP  ----> Escojemos cunato se aleja el grafico del borde superior (Se mide en puntos)
        # 4- WIDTH ---> (Se mide en puntos)
            # -1 para conservar el ancho del archivo existente
        # 5- HEIGHT ---> (Se mide en puntos)
        # Este es el nombre de la foto que vamos a utilizar en el front end para llamar a la imagen
        if self.pic == 'True':
            path = self.AB_PIC_PATH + '\\'+ 'wk.png'
        else:
            path = self.AB_PIC_PATH + '\\'+ self.cName+'.png'


        x = self.type_graphic

        if x == 'xl3DColumn':
            # HAcemos esto para que el grafico se ajuste al numero de columnas.
            if (lastcol-1) == 4:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,530,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,990, 105,-1,-1)
            elif (lastcol-1) == 5:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,660,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1115, 105,-1,-1)
            elif (lastcol-1) == 6:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,830,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1290, 105,-1,-1)
            else:
                # Este es el que tengo por defecto cunsdo no selecionan columnas
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,350,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,825, 105,-1,-1)

        elif x == 'xlCylinderCol':
            # HAcemos esto para que el grafico se ajuste al numero de columnas.
            if (lastcol-1) == 4:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,530,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,990, 105,-1,-1)
            elif (lastcol-1) == 5:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,660,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1115, 105,-1,-1)
            elif (lastcol-1) == 6:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,830,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1290, 105,-1,-1)
            else:
                # Este es el que tengo por defecto cunsdo no selecionan columnas
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,350,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,825, 105,-1,-1)

        elif x == 'xl3DLine':
            # HAcemos esto para que el grafico se ajuste al numero de columnas.
            if (lastcol-1) == 4:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,530,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,990, 105,-1,-1)
            elif (lastcol-1) == 5:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,660,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1115, 105,-1,-1)
            elif (lastcol-1) == 6:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,830,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,1290, 105,-1,-1)
            else:
                # Este es el que tengo por defecto cunsdo no selecionan columnas
                Sheet3.Shapes.AddChart(win32c.xl3DLine,350,100,600,350).Select()
                Sheet3.Shapes.AddPicture(path,True,True,825, 105,-1,-1)

        self.Excel.ActiveChart.SetSourceData(Source=Sheet3.Range('A7:'+letter+str(lastrow)))

    def final(self):
        # open workbooks
        f1_name = '\\1.xlsx'
        f2_name = '\\2.xlsx'
        #f3_name = '\\3.xlsx'
        f5_name = '\\5.xlsx'

        filename1 = self.AB_FINAL_SAVE_PATH + f1_name
        filename2 = self.AB_FINAL_SAVE_PATH + f2_name
        #filename3 = self.AB_FINAL_SAVE_PATH + f3_name
        filename5 = self.AB_FINAL_SAVE_PATH + f5_name


        wb1 = self.Excel.Workbooks.Open(filename1)
        wb2 = self.Excel.Workbooks.Open(filename2)
        #wb3 = self.Excel.Workbooks.Open(filename3)
        wb5 = self.Excel.Workbooks.Open(filename5)

        sheetname = "Monthly Trend"
        sheetname1 = "Monthly_Trend"

        # copy worksheet above from wb1 to wb2
        #wb1.Worksheets(sheetname).Copy(Before=wb2.Sheets(1))
        # Esto es para que lleve el orden normal en que entregamos los reportes
        wb5.Worksheets(sheetname1).Copy(After=wb2.Sheets(1))

        if self.Monthly_Trend1 == 'Monthly_Trend1':
            #---- CAMBIO EL NOMBRE FINAL DEL TAB ------------------
            # Puedo cambiarle el nombre aqui al final antes de guardar.
            wb1.Worksheets(sheetname).Name = ".."
            wb1.Worksheets('..').Copy(After=wb2.Sheets(2))
        else:
            pass
        #wb3.Worksheets(sheetname1).Copy(After=wb2.Sheets(2))

        #------------ YEAR COMPARISON -------------------
        if self.Yearly_Comparison == 'Yearly_Comparison':

            f6_name = '\\6.xlsx'
            filename6 = self.AB_FINAL_SAVE_PATH + f6_name
            wb6 = self.Excel.Workbooks.Open(filename6)
            sheetname = "Yearly_Comparison"

            #---- CAMBIO EL NOMBRE FINAL DEL TAB ------------------
            # Puedo cambiarle el nombre aqui al final antes de guardar.
            wb6.Worksheets(sheetname).Name = "Yearly_Comparison"
            wb6.Worksheets('Yearly_Comparison').Copy(After=wb2.Sheets(2))
            wb6.Save()
        else:
            pass
        #------- SET DEFAULT TAB ---------------------------
        # Esto es para que el tab quede como DEFAULT cuando se abra la spreadsheet
        wb2.Worksheets('Book Stats').Activate()

        # Aqui tenemos que salvar porque sino me queda abierto y se sale una ventana solicitanto si quiero salvarlo.
        wb1.Save()
        wb2.Save()
        #wb3.Save()
        wb5.Save()

        # --------- REMOVE IF THERE IS A FILE ALREADY WITH THIS FILE
        if os.path.isfile(self.AB_FINAL_SAVE_PATH_PERMANENT + '\\' + self.cName + '_' + 'Book_Stats' + '_' + self.initMonth + self.startYear +'-'+ self.endMonth + self.closeYear +'.xlsx'):
            os.remove(self.AB_FINAL_SAVE_PATH_PERMANENT + '\\' + self.cName + '_' + 'Book_Stats' + '_'+ self.initMonth + self.startYear +'-'+ self.endMonth + self.closeYear +'.xlsx')

        # Ahora guardo el nuevo archivo en la ruta final
        wb2.SaveAs(self.AB_FINAL_SAVE_PATH_PERMANENT + '\\'+ self.cName + '_' + 'Book_Stats' + '_'+ self.initMonth + self.startYear +'-'+ self.endMonth + self.closeYear +'.xlsx')

        # Salvo en la ruta temporal.
        wb2.SaveAs(self.AB_FINAL_SAVE_PATH + '\\'+ self.cName + '_' + 'Book_Stats' + '_'+ self.initMonth + self.startYear +'-'+ self.endMonth + self.closeYear +'.xls')

        self.Excel.Quit()

        institution_name = self.cName
        institution_init_year = self.startYear
        institution_end_year= self.closeYear

        find_path = str(self.cName + '_' + 'Book_Stats' + '_' + self.initMonth + self.startYear +'-'+ self.endMonth + self.closeYear +'.xlsx')

        type = 'Books'

        #-------------- SELECT IMAGE -----------------------------------
        # Este es el nombre de la foto que vamos a utilizar en el front end para llamar a la imagen
        if self.pic == 'True':
            pic_path =  'wk.png'
        else:
            pic_path =  str(self.cName+'.png')

        # -------------  BORRAR archivos ------------------
        #Borro los archivos de 1 y 2
        list = [f1_name, f2_name]
        for i in list:
            if os.path.isfile(i):
                os.remove(i)
        else:
            print("The system cannot find the file specified")

        return institution_name, institution_init_year, institution_end_year,find_path, type, pic_path

    def clean_folder(self):

        path = self.TEMP_SAVE_PATH

        # Esto es para terminar de borrar los archivos que quedan en el folder principal.
        path1 = self.FINAL_SAVE_PATH

        for file_name in os.listdir(path):
            # construct full file path
            file = path + file_name
            if os.path.isfile(file):
                os.remove(file)

        # Ahora necesito remover los otros documentos del UPLOADS directory
        for file_name in os.listdir(path1):
            # construct full file path
            file = path1 + file_name
            if os.path.isfile(file):
                os.remove(file)

#------------------------------------------------------------------------------
                #---------- MONTLY TREND- --------------#
#------------------------------------------------------------------------------
    def extra(self):

        source_file = self.AB_FINAL_SAVE_PATH + '\\1.xlsx'

        self.Excel.Visible = False
        #-------- OPEN Book ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = self.Excel.Workbooks.Open(source_file)

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        ws = wb.Sheets.Add()
        ws.Name = "Monthly_Trend"

        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el , porque queremos asignar la hoja dos que acabamos de crear.
        Sheet3 = wb.Worksheets('Monthly_Trend')


        # Esta es la forma de seleccionar el SHEET dentro de work book.
        Sheet4 = wb.Worksheets('Monthly Trend')

        #------------------------------
        # -------- LAST (COL - ROW) ----
        # ------------------------------
        # Esto nos sirve para averiguar cual es el ultimo campo utilizado en la tabla.
        lastrow = Sheet4.UsedRange.Rows.Count
        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet4.UsedRange.Columns.Count
        # Asi convierto el numero de columna en character
        # Le sumo 1 porque la primera columna no es parte de los resultados .
        # SI la dejo asi me cuenta una columna menos
        letter = chr(ord('@') + int(lastcol)+1)

        # copy worksheet above from wb1 to wb2
        #wb1.Worksheets(sheetname).Copy(Before=wb2.Sheets(1))
        # Esto es para que lleve el orden normal en que entregamos los reportes
        #wb1.Worksheets(sheetname).Copy(After=wb2.Sheets(1))


        # Le doy un 2+size para que no me deje mada botado.
        #https://learn.microsoft.com/en-us/previous-versions/office/developer/office-2007/bb178833(v=office.12)?redirectedfrom=MSDN
        Sheet4.Range("A7:"+letter+str(2+lastrow)).SpecialCells(12).Copy(Sheet3.Range('A7:'+letter+str(2+lastrow)))

        #Sheet3.Range('A7:C'+str(lastrow)).Table()
#---------------------------------------------------------------------------------------

        #Sheet3 = Book.set_table_extra(self, Sheet3)

        Book.customize_title(self,Sheet3)

        #Database.extra_chart(self,Sheet3)

        #-------- SAVE WORKBOOK
        extra_path = self.AB_FINAL_SAVE_PATH + '\\3.xlsx'
        wb.SaveAs(self.AB_FINAL_SAVE_PATH + '\\3.xlsx')

        time.sleep(2)
        #-------- QUIT EXCEL
        self.Excel.Application.Quit()

        col_end = str(letter)

        return extra_path, col_end

#---------- SET UP THE TABLE FORMAT- --------------
    def set_table_extra(self, extra_path, col_end):
        #---------------------#
        #       WORK BOOK     #
        #---------------------#
        #-------- OPEN BOOK ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = load_workbook(extra_path)
        Sheet = wb['Monthly_Trend']


        #------  FIND SIZE DF --------------------
        # Podemos averiguar el tamaño de la columna facilmente con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(extra_path, engine='openpyxl')

        rows = df.shape[0]
        cols = df.shape[1]

        letter = chr(ord('@') + int(cols))
        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))
        ######################
        #  CREATE TABLE RANGE
        ######################

        # Creamos un objeto de data que guarde los datos de la tabla
        # el parametro 'ref' indica el rango de los datos
        # Necesitamos cojer desde la columna A hasta la M--- porque el el numero de columnas que dejamos
        # A esto le agregamos el numero de rows --SIZE, de esta forma cubrimos todos los datos dentro de la tabla

        tab = Table(displayName='Table10',ref = 'A7:'+letter+size)

        ##################
        #  STYLE TO TABLE
        ##################
        # Ahora le damos estilo a la tabla
        # Creo una variable que guarde el objeto de estilo
        # Solo quiero que los rows tengan strpes. Las columnas las mantengo normales
        # SHOWFIRST-SHOWLAST ----Sirve para aplicar el formato de la table a todas las columnas o solo a algunas.
        style = TableStyleInfo(name='TableStyleLight1', showFirstColumn=False, showLastColumn= False,
                                                            showRowStripes=True, showColumnStripes=False)
        # ahora asigno la variable
        tab.tableStyleInfo = style

        # Por ultimo agregamos la tabla al worksheet
        Sheet.add_table(tab)

        #-*---------------------------------------------
        #--------- ADJUST WIDTH -------------------------
        #--------------------------------------------------
        # Ajustamos las cokumnas de manera diferente porque estamos utilizando Open XL library
        # Esto es otr forma de hacerlo
        #from openpyxl.utils import get_column_letter

        # COl


        Sheet.column_dimensions['A'].width = 20
        Sheet.column_dimensions['B'].width = 20
        Sheet.column_dimensions['C'].width = 30
        Sheet.column_dimensions['D'].width = 30
        Sheet.column_dimensions['E'].width = 30
        Sheet.column_dimensions['F'].width = 30


        path = self.AB_FINAL_SAVE_PATH + '\\4.xlsx'
        # path = self.AB_PIC_PATH + '\\4.xlsx'
        wb.save(path)
        #wb1.SaveAs(path)

        Book.extra_chart(self,path)

    def extra_chart(self,path):

        self.Excel.Visible = False
        #-------- OPEN Book ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = self.Excel.Workbooks.Open(path)

        # Esta es la forma de seleccionar el SHEET dentro de work book.
        Sheet3 = wb.Worksheets('Monthly_Trend')

        #------  FIND SIZE DF --------------------
        # Podemos averiguar el tamaño de la columna facilmente con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(path, engine='openpyxl')
        rows = df.shape[0]
        cols = df.shape[1]

        letter = chr(ord('@') + int(cols))

        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        size1 = int(size) - 1

        Sheet3.Range('A7:'+letter+str(size1)).Select()

        #------------------------#
        #        CUSTOMIZE TITLE
        #------------------------#
        # ------- Merge cells
        # Escojo el rango de cells
        # El primer numero es rows# Segundo numero es column
        #myRange = Sheet3.Range(Sheet3.Cells(3,1),Sheet3.Cells(3,cols))
        #myRange.MergeCells = True
        # Las options de align son : Center, Right,Left
        #myRange.HorizontalAlignment = win32c.xlLeft
        # -------How to add colors
        #https://itecnote.com/tecnote/python-pywin32-excel-formatting/
        # myRange.Interior.ColorIndex = 6
        #Shet3.Range("A1", "D1").Interior.ColorIndex = 6
        cname1 = self.cName.replace('_', ' ')
        Sheet3.Cells(3,1).Value = cname1 + ' ' + 'Book Stats' + ' '+ '  ('+ self.initMonth +' '+self.startYear +' - '+' '+ self.endMonth + ' '+self.closeYear+') '
        #---- Customize title
        Sheet3.Cells(3,1).Font.Size = 15
        Sheet3.Cells(3,1).Font.Bold = True
        #https://www.w3schools.com/colors/colors_picker.asp
        Sheet3.Cells(3,1).Font.Color = Book.RGB(0, 0, 0)

        # --------- HIDE Grind --------------

        if self.use_grid == 'True':
            self.Excel.ActiveWindow.DisplayGridlines = True
        else:
            self.Excel.ActiveWindow.DisplayGridlines = False

        # --------------------------------------
        #  SELECT TYPE OF GRAPH
        # --------------------------------------
        # Al opcion ADDCHART ofrece muchas posibilidades de customizacion
        # 1- Tipo de grafico-----> solo tenemos que decidir el tipo de grafico que deseamos utilizar
            ##https://docs.microsoft.com/es-es/office/vba/api/excel.xlcharttype
                    #Excel.ActiveChart.ChartType = win32c.xlLine
        # 2- LEFT------> Escojemos cuanto queremos que se aleje el grafico del borde izquierdo (Se mide en puntos)
        # 3- TOP  ----> Escojemos cunato se aleja el grafico del borde superior (Se mide en puntos)
        # 4- WIDTH ---> (Se mide en puntos)
        # 5- HEIGHT ---> (Se mide en puntos)

        ###############################################################
        #------------ IMAGE ----------------------
        ###############################################################
        # https://learn.microsoft.com/es-es/office/vba/api/excel.shapes.addpicture
        # 1- Path-----> definimos donde está el archivo
                        # Eventualmente será necesario hacer una base de datos referencial de imagenes.
        # TRUE --- vincular la imagen al archivo desde el que se creó.

        # True --------> guardar la imagen vinculada con el documento en el que se inserta.
        # 2- LEFT------> Escojemos cuanto queremos que se aleje el grafico del borde izquierdo (Se mide en puntos)
        # 3- TOP  ----> Escojemos cunato se aleja el grafico del borde superior (Se mide en puntos)
        # 4- WIDTH ---> (Se mide en puntos)
            # -1 para conservar el ancho del archivo existente
        # 5- HEIGHT ---> (Se mide en puntos)

        x = self.type_graphic

        #----------------------------------------
        #   Select imagen
        #----------------------------------------

        # Primero tenemos que definir cual es la imagen que vamos a utilizar
        # Este es el nombre de la foto que vamos a utilizar en el front end para llamar a la imagen
        if self.pic == 'True':
            image= self.AB_PIC_PATH + '\\'+ 'wk.png'
        else:
            image = self.AB_PIC_PATH + '\\'+ self.cName+'.png'

        #----------------------------------------
        #   Select Chart
        #----------------------------------------
        if x == 'xl3DColumn':
            # HAcemos esto para que el grafico se ajuste al numero de columnas.
            if cols == 3:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,420,90,600,350).Select()

                # -> FOTO -----475
                # -> STEP -----165
                Sheet3.Shapes.AddPicture(image,True,True,895, 105,-1,-1)
            elif cols == 4:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,585,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1060, 105,-1,-1)
            elif cols == 5:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,750,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1225, 105,-1,-1)
            else:
                Sheet3.Shapes.AddChart(win32c.xl3DColumn,915,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1390, 105,-1,-1)

        elif x == 'xlCylinderCol':

            if cols == 3:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,400,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,875, 105,-1,-1)
            elif cols == 4:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,500,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,975, 105,-1,-1)
            elif cols == 5:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,600,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1075, 105,-1,-1)
            else:
                Sheet3.Shapes.AddChart(win32c.xlCylinderCol,700,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1175, 105,-1,-1)

        elif x == 'xl3DLine':

            if cols == 3:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,400,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,875, 105,-1,-1)
            elif cols == 4:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,500,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,975, 105,-1,-1)
            elif cols == 5:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,600,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1075, 105,-1,-1)
            else:
                Sheet3.Shapes.AddChart(win32c.xl3DLine,700,90,600,350).Select()
                Sheet3.Shapes.AddPicture(image,True,True,1175, 105,-1,-1)
        else:
            Sheet3.Shapes.AddChart(win32c.xlLine,800,90,600,350).Select()
            Sheet3.Shapes.AddPicture(image,True,True,1175, 105,-1,-1)

        ## Aqui utilizo el SIZE1 porque no quiero incluir los totales en la grafica.
        # Entonces para evitarlo quito el ultimo row que son los TOTALS
        self.Excel.ActiveChart.SetSourceData(Source=Sheet3.Range('A7:'+letter+str(size1)))


        wb.SaveAs(self.AB_FINAL_SAVE_PATH + '\\5.xlsx')

        self.Excel.Application.Quit()

    def set_image(self):
        # Creamos la variable para encontrar la picture
        path = self.AB_PIC_PATH + '\\'+ self.cName+'.png'
        # Si existe la imagen no devuelve TRUE sino nos devuelve FALSE
        isExist = os.path.isfile(path)
        return isExist

#------------------------------------------------------------------------------
                #---------- Yearly comparison- --------------#
#------------------------------------------------------------------------------
    def yearly_comp(self):

        source_file = self.AB_TEMP_SAVE_PATH + '\\step_3.xlsx'

        self.Excel.Visible = False
        #-------- OPEN Book ------------------
        # Primero abro el libro que ya tiene la columna incluida
        wb = self.Excel.Workbooks.Open(source_file)

        # Esta es la forma de seleccionar el SHEET dentro de work book.
        ws = wb.Worksheets('.')

        #------  FIND SIZE DF --------------------
        # Podemos averiguar el tamaño de la columna facilmente con pandas
        # Llamo a un dataframe para buscar la info rapido y asigno una variable
        df = pd.read_excel(source_file, engine='openpyxl')
        rows = df.shape[0]
        cols = df.shape[1]
        # Aqui buscamos el tamaño de la table_test
        # Agregamos 1 para que cubra el ultimo row de abajo
        # Lo pasamos a string para poder pegarlo en una cadena
        size = str(1+(df.shape[0]))

        #------  SELECT RANGE  --------------------
        # Pongo O porque se que tengo 15 columnas que son la que deje despues de filtrar la informacion
        # Como el numero de datos puede variar necesito hacerlo dinamico.
        PivotSourceRange = ws.Range('A1:Q'+size)

        #############################################################################################################
        #                           FIRST PIVOT TABLE
        #############################################################################################################
        #-------- CREATE PIVOT TABLE
        # Aqui le pasamos los rangos a una variable
        #PivotSourceRange = wb.Range('A1:M'+size)

        #---HACK ---> GOOD STUFF
        # Si creamos la segunda hoja al principio del documento no da error en el paso anterior
        # Descubri que es necesario creala justo despues que ya hemos seleccionado la data desde la primera hoja
        hoja3 = wb.Sheets.Add()
        hoja3.Name = "Yearly_Comparison"

        # Creamos una variable que almacene esta data
        # Como parametro le pasamos el 1, porque queremos asignar la hoja dos que acabamos de crear.
        Sheet3 = wb.Worksheets(1)

        # --------- HIDE Grind --------------

        if self.use_grid == 'True':
            self.Excel.ActiveWindow.DisplayGridlines = True
        else:
            self.Excel.ActiveWindow.DisplayGridlines = False
        # --------- HIDE Grind --------------

        ############ CHECK POINT------> Hasta aqui todo bien

        # Una vez seleccionada la data le indicamos en donde deseamos colocar el pivot_table
        # Según conversacones con Jeff comenzamos en el row 7 de la primera columna.
        cl3=Sheet3.Cells(7,1)
        #------------------------#
        #        Title of Pivot
        #------------------------#
        Book.customize_title(self,Sheet3)

        # Ahora insertamos esos datos dentro del rango seleccionado.
        PivotTargetRange=  Sheet3.Range(cl3,cl3)
        PivotTableName="ReportPivotTable"

        # Esto aún no lo entiendo bien pero es la forma de inicializar la pivot table
        PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)

        PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

        #-----------------#
        #        ROW
        #-----------------#
        PivotTable.PivotFields('Year').Orientation = win32c.xlRowField
        PivotTable.PivotFields('Year').Position = 1
        #-----------------#
        #        FILTER
        #-----------------#
        PivotTable.PivotFields('Title').Orientation = win32c.xlPageField
        PivotTable.PivotFields('Title').Position = 1
        # Esto es en caso que querramos tener algo seleccionado por default
        #PivotTable.PivotFields('Gender').CurrentPage="M"
        #-----------------#
        #        VALUES (Columnas)
        #-----------------#

        # Auiq lo que hago es crear un array para guardar todas las opciones que han sido seleccinadas
        # Evaluo la condicion uno por uno y si existe lo voy metiendo en el array para trabajar el STYLE
        selection_list = []

        if  self.session_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Sessions'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Sessions')
        else:
            pass

        if self.TII_select:
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Investigations'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Investigations')
        else:
            pass

        if self.UII_select == 'Unique_Item_Investigations':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Investigations'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Investigations')
        else:
            pass

        if self.TIR_select == 'Total_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Total_Item_Requests'))
            DataField.NumberFormat="#" ##0.00'
            selection_list.append('Total_Item_Requests')
        else:
            pass

        if self.UIR_select == 'Unique_Item_Requests':
            DataField = PivotTable.AddDataField(PivotTable.PivotFields('Unique_Item_Requests'))
            DataField.NumberFormat="#"##0.00'
            selection_list.append('Unique_Item_Requests')
        else:
            pass


        #-----------------#
        #       FORMAT
        #-----------------#
        # Esto es para que aparezca striped
        PivotTable.ShowTableStyleRowStripes = True

        #--------------------------#
        #       CHANGE FIELD NAMES
        #-------------------------#
        arg1 = 'Year'
        arg4 = 'Book'

        # Utilizo el metodo style para cambiar el  Estilo y nombre de los campos
        Book.style(self, Sheet3, arg1, selection_list, arg4, size)
        #-------- SAVE WORKBOOK
        #Sheet3.Columns.AutoFit()

        #--------------------------#
        #       CRETE CHART
        #-------------------------#
        Book.create_chart(self, Sheet3)

        #--------------------------#
        #       MEAN AND TREND
        #-------------------------#

        list, sessions, TIR, UIR, TII, UII = Book.average(self, Sheet3)

        Book.trend(self, list, sessions, TIR, UIR, TII, UII,Sheet3)

        #---------------------------------
        #
        #---------------------------------

        #wb.SaveAs(self.AB_FINAL_SAVE_PATH + self.cName + ' ' + 'Database_Stats' + ' '+ self.startYear +'-'+ self.closeYear +'.xlsx')
        wb.SaveAs(self.AB_FINAL_SAVE_PATH + '\\6.xlsx')

        time.sleep(2)
        #-------- QUIT EXCEL
        self.Excel.Application.Quit()


    def average(self, Sheet3):

        df = pd.read_excel(self.SOURCE_FILE)

        # Averiguar el tamaño de cada grupo
        x = df.groupby("Year").size()

        # Para saber cuantos grupos hay
        # En este caso para saber cuales años tenemos.
        # Es decir las clases de esta columna
        year = df.Year.nunique()
        #---------------------------------
        #         Inicializo Arrays
        #---------------------------------

        sessions_2019 = []
        sessions_2020 = []
        sessions_2021 = []
        sessions_2022 = []

        TII_2019 = []
        TII_2020 = []
        TII_2021 = []
        TII_2022 = []

        UII_2019 = []
        UII_2020 = []
        UII_2021 = []
        UII_2022 = []

        TIR_2019 = []
        TIR_2020 = []
        TIR_2021 = []
        TIR_2022 = []

        UIR_2019 = []
        UIR_2020 = []
        UIR_2021 = []
        UIR_2022 = []

        #---------------------------------
        #         Clasifico dataset
        #---------------------------------
        # Dependinedo del año meto en el array TODOS los parametros que se necesitan
        # Meto cada columna en una variable y las voy acumulando en el array con un nombre.
        for index, row in df.iterrows():
            if row['Year'] == 2019:
                sessions_2019.append(row['Sessions'])
                TIR_2019.append(row['Total_Item_Requests'])
                UIR_2019.append(row['Unique_Item_Requests'])
                UII_2019.append(row['Unique_Item_Investigations'])
                TII_2019.append(row['Total_Item_Investigations'])

            elif row['Year'] == 2020:
                sessions_2020.append(row['Sessions'])
                TIR_2020.append(row['Total_Item_Requests'])
                UIR_2020.append(row['Unique_Item_Requests'])
                UII_2020.append(row['Unique_Item_Investigations'])
                TII_2020.append(row['Total_Item_Investigations'])

            elif row['Year'] == 2021:
                sessions_2021.append(row['Sessions'])
                TIR_2021.append(row['Total_Item_Requests'])
                UIR_2021.append(row['Unique_Item_Requests'])
                UII_2021.append(row['Unique_Item_Investigations'])
                TII_2021.append(row['Total_Item_Investigations'])

            elif row['Year'] == 2022:
                sessions_2022.append(row['Sessions'])
                TIR_2022.append(row['Total_Item_Requests'])
                UIR_2022.append(row['Unique_Item_Requests'])
                UII_2022.append(row['Unique_Item_Investigations'])
                TII_2022.append(row['Total_Item_Investigations'])

            else:
                pass

        #---------------------------------
        #         AVERAGE
        #---------------------------------

        #------------Inicio Arrays----------------
        # Aqui guardo las operaciones de cada columna
        avg_sessions_list = []
        avg_TIR_list = []
        avg_UIR_list = []
        avg_UII_list = []
        avg_TII_list = []

        #------------SESSIONS----------------
        #Saco el average por cada año
        avg_sessions_2019 = sum(sessions_2019)/year
        avg_sessions_2020 = sum(sessions_2020)/year
        avg_sessions_2021 = sum(sessions_2021)/year
        avg_sessions_2022 = sum(sessions_2022)/year

        # Lo meto dentro de la lista
        avg_sessions_list.append(avg_sessions_2019)
        avg_sessions_list.append(avg_sessions_2020)
        avg_sessions_list.append(avg_sessions_2021)
        avg_sessions_list.append(avg_sessions_2022)

        avg_sessions = round(sum(avg_sessions_list))

        #------------TIR--------------

        avg_TIR_2019 = sum(TIR_2019)/year
        avg_TIR_2020 = sum(TIR_2020)/year
        avg_TIR_2021 = sum(TIR_2021)/year
        avg_TIR_2022 = sum(TIR_2022)/year

        avg_TIR_list.append(avg_TIR_2019)
        avg_TIR_list.append(avg_TIR_2020)
        avg_TIR_list.append(avg_TIR_2021)
        avg_TIR_list.append(avg_TIR_2022)

        avg_TIR = round(sum(avg_TIR_list))

        #------------UIR--------------

        avg_UIR_2019 = sum(UIR_2019)/year
        avg_UIR_2020 = sum(UIR_2020)/year
        avg_UIR_2021 = sum(UIR_2021)/year
        avg_UIR_2022 = sum(UIR_2022)/year

        avg_UIR_list.append(avg_UIR_2019)
        avg_UIR_list.append(avg_UIR_2020)
        avg_UIR_list.append(avg_UIR_2021)
        avg_UIR_list.append(avg_UIR_2022)

        avg_UIR = round(sum(avg_UIR_list))

        #------------UII--------------

        avg_UII_2019 = sum(UII_2019)/year
        avg_UII_2020 = sum(UII_2020)/year
        avg_UII_2021 = sum(UII_2021)/year
        avg_UII_2022 = sum(UII_2022)/year

        avg_UII_list.append(avg_UII_2019)
        avg_UII_list.append(avg_UII_2020)
        avg_UII_list.append(avg_UII_2021)
        avg_UII_list.append(avg_UII_2022)

        avg_UII = round(sum(avg_UII_list))

        #------------TII--------------

        avg_TII_2019 = sum(TII_2019)/year
        avg_TII_2020 = sum(TII_2020)/year
        avg_TII_2021 = sum(TII_2021)/year
        avg_TII_2022 = sum(TII_2022)/year

        avg_TII_list.append(avg_TII_2019)
        avg_TII_list.append(avg_TII_2020)
        avg_TII_list.append(avg_TII_2021)
        avg_TII_list.append(avg_TII_2022)

        avg_TII = round(sum(avg_TII_list))

        global final
        # Acumulo los resultados de cada columna(a su vez contine los de cada año) y envio el array por un return
        final = [avg_sessions, avg_TIR, avg_UIR,  avg_TII, avg_UII]

        #---------------------------------
        #         Trend
        #---------------------------------
        # Esta informacion es la que envio a la funcion de trend para que me haga las operaciones
        # Lo hago asi porque sino el codigo se me hace muy grande.
        # Lo exporto como diccionario para poder buscar por KEY
        sessions = {'2019':sessions_2019,'2020':sessions_2020,'2021':sessions_2021,'2022':sessions_2022}

        TIR = {'2019':TIR_2019,'2020':TIR_2020,'2021':TIR_2021,'2022':TIR_2022}
        UIR = {'2019':UIR_2019, '2020':UIR_2020, '2021':UIR_2021,'2022':UIR_2022}
        UII = {'2019':UII_2019,'2020':UII_2020,'2021':UII_2021,'2022':UII_2022}
        TII = {'2019':TII_2019,'2020':TII_2020,'2021':TII_2021,'2022':TII_2022}


        # https://www.automateexcel.com/vba/find-last-row-column-cell/
        lastrow = Sheet3.UsedRange.Rows.Count

        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet3.UsedRange.Columns.Count

        lastcol1 = lastcol-1

        # Asi convierto el numero de columna en character
        letter = chr(ord('@') + int(lastcol1))

        # ------- Merge cells
        # Escojo el rango de cells
        # El primer numero es rows# Segundo numero es column
        #myRange = Sheet3.Range(Sheet3.Cells((lastrow+3),1),Sheet3.Cells((lastrow+3),5))
        #myRange.MergeCells = True
        # Las options de align son : Center, Right,Left
        #myRange.HorizontalAlignment = win32c.xlLeft
        # -------How to add colors
        #https://itecnote.com/tecnote/python-pywin32-excel-formatting/
        #myRange.Interior.ColorIndex = 6
        #Shet3.Range("A1", "D1").Interior.ColorIndex = 6

        # ----------------------------------------------------------------
        #           AVERAGE
        # ----------------------------------------------------------------


        Sheet3.Cells((lastrow+1),1).Value = 'Average'


        # ----------------------------------------------------------------
        #           EDGES
        # ----------------------------------------------------------------
        #https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)

        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlInsideHorizontal).Color = Book.RGB(0, 0, 0)
        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Borders(win32c.xlInsideVertical).Color = Book.RGB(0, 0, 0)



        Sheet3.Cells((lastrow+1),1).Font.Size = 13
        Sheet3.Cells((lastrow+1),1).Font.Bold = True
        Sheet3.Cells((lastrow+1),1).Font.Color = Book.RGB(0,0,0)
        Sheet3.Cells((lastrow+1),1).HorizontalAlignment = win32c.xlCenter
        #Sheet3.Range('A'+str(lastrow+4),letter+str(lastcol)).Interior.ColorIndex = 6

        Sheet3.Range('A'+str(lastrow+1),letter+str(lastrow+1)).Interior.ColorIndex = 6

        # ----------------------------------------------------------------
        #           SHOW VALUES
        # ----------------------------------------------------------------

        if  self.session_select:
            Sheet3.Cells((lastrow+1),2).Value = final[0]
        else:
            pass

        if self.TII_select:
            Sheet3.Cells((lastrow+1),3).Value = final[3]
        else:
            pass

        if self.UII_select ==  'Unique_Item_Investigations':
            Sheet3.Cells((lastrow+1),4).Value = final[4]
        else:
            pass

        if self.TIR_select == 'Total_Item_Requests':
            Sheet3.Cells((lastrow+1),5).Value = final[1]
        else:
            pass


        if self.UIR_select == 'Unique_Item_Requests':
            Sheet3.Cells((lastrow+1),6).Value = final[2]
        else:
            pass

        # Una vez que tenemos toda la infomracion la regresamos a la funcion.
        return final, sessions, TIR, UIR, TII, UII

    def trend(self, list, sessions, TIR, UIR, TII, UII,Sheet3):

        # Aqui incluye el merge del titulo
        # Necesito quitarle algunas columnas
        lastcol = Sheet3.UsedRange.Columns.Count

        # Asi convierto el numero de columna en character
        letter = chr(ord('@') + int((lastcol)-2))

        lastrow = Sheet3.UsedRange.Rows.Count
        # Inicializo lo array donde gurado los resultados del trend
        # Tengo que inicializar el array para cada uno de las variables que voy a utilizar.
        # Es una para cada año por variable
        #-----------------------
        sessions_trend_2019 = []
        sessions_trend_2120 = []
        sessions_trend_2221 = []
        #----------------------
        #-----------------------
        TIR_trend_2019 = []
        TIR_trend_2120 = []
        TIR_trend_2221 = []
        #-----------------------
        UIR_trend_2019 = []
        UIR_trend_2120 = []
        UIR_trend_2221 = []
        #-----------------------
        UII_trend_2019 = []
        UII_trend_2120 = []
        UII_trend_2221 = []

        #-----------------------
        TII_trend_2019 = []
        TII_trend_2120 = []
        TII_trend_2221 = []
#---------------------------------
#         Recorro diccionarios
#---------------------------------

        #---------------------------------
        #         SESSIONS
        #---------------------------------
        for key, value in sessions.items():
# -------  Trend 2020 -2019
            if key == '2019':
                # Siempre se compara con el año anterior para saber la tendencia
                if sum(sessions[key]) > 0:
                    sessions_trend_2019 = sum(sessions['2020'])/sum(sessions[key])-1
                else:
                    # Si al final resulta 0 entonces le doy un 1 para que no me de el errores
                    # Despues tengo que pasarlo a porcentaje y me daria 100%. Esto significa que no existen datos
                    sessions_trend_2019 = 1

# -------  Trend 2021 - 2020
            elif key == '2020':
                if sum(sessions[key]) > 0:
                    sessions_trend_2120 = sum(sessions['2021'])/sum(sessions[key])-1
                else:
                    sessions_trend_2120 = 1

# -------  Trend 2022 - 2021
            elif key == '2021':
                if sum(sessions[key]) > 0:
                    sessions_trend_2221 = sum(sessions['2022'])/sum(sessions[key])-1
                else:
                    sessions_trend_2221 = 1

            else:
                pass

        #---------------------------------
        #         TOtal Item requests
        #---------------------------------
        for key, value in TIR.items():
# -------  Trend 2020 -2019
            if key == '2019':
                # Siempre se compara con el año anterior para saber la tendencia
                if sum(TIR[key]) > 0:
                    TIR_trend_2019 = sum(TIR['2020'])/sum(TIR[key])-1
                else:
                    # Si al final resulta 0 entonces le doy un 1 para que no me de el errores
                    # Despues tengo que pasarlo a porcentaje y me daria 100%. Esto significa que no existen datos
                    TIR_trend_2019 = 1

# -------  Trend 2021 - 2020
            elif key == '2020':
                if sum(TIR[key]) > 0:
                    TIR_trend_2120 = sum(TIR['2021'])/sum(TIR[key])-1
                else:
                    TIR_trend_2120 = 1

# -------  Trend 2022 - 2021
            elif key == '2021':
                if sum(TIR[key]) > 0:
                    TIR_trend_2221 = sum(TIR['2022'])/sum(TIR[key])-1
                else:
                    TIR_trend_2221 = 1

            else:
                pass
        #---------------------------------
        #         Unique Item requests
        #---------------------------------
        for key, value in UIR.items():
# -------  Trend 2020 -2019
            if key == '2019':
                # Siempre se compara con el año anterior para saber la tendencia
                if sum(UIR[key]) > 0:
                    UIR_trend_2019 = sum(UIR['2020'])/sum(UIR[key])-1
                else:
                    # Si al final resulta 0 entonces le doy un 1 para que no me de el errores
                    # Despues tengo que pasarlo a porcentaje y me daria 100%. Esto significa que no existen datos
                    UIR_trend_2019 = 1

# -------  Trend 2021 - 2020
            elif key == '2020':
                if sum(UIR[key]) > 0:
                    UIR_trend_2120 = sum(UIR['2021'])/sum(UIR[key])-1
                else:
                    UIR_trend_2120 = 1

# -------  Trend 2022 - 2021
            elif key == '2021':
                if sum(UIR[key]) > 0:
                    UIR_trend_2221 = sum(UIR['2022'])/sum(UIR[key])-1
                else:
                    UIR_trend_2221 = 1

            else:
                pass
        #---------------------------------
        #         Unique Item Investigfation
        #---------------------------------
        for key, value in UII.items():
# -------  Trend 2020 -2019
            if key == '2019':
                # Siempre se compara con el año anterior para saber la tendencia
                if sum(UII[key]) > 0:
                    UII_trend_2019 = sum(UII['2020'])/sum(UII[key])-1
                else:
                    # Si al final resulta 0 entonces le doy un 1 para que no me de el errores
                    # Despues tengo que pasarlo a porcentaje y me daria 100%. Esto significa que no existen datos
                    UII_trend_2019 = 1

# -------  Trend 2021 - 2020
            elif key == '2020':
                if sum(UII[key]) > 0:
                    UII_trend_2120 = sum(UII['2021'])/sum(UII[key])-1
                else:
                    UII_trend_2120 = 1

# -------  Trend 2022 - 2021
            elif key == '2021':
                if sum(UII[key]) > 0:
                    UII_trend_2221 = sum(UII['2022'])/sum(UII[key])-1
                else:
                    UII_trend_2221 = 1

            else:
                pass

        #---------------------------------
        #         Total Item Investigfation
        #---------------------------------
        for key, value in TII.items():
# -------  Trend 2020 -2019
            if key == '2019':
                # Siempre se compara con el año anterior para saber la tendencia
                if sum(TII[key]) > 0:
                    TII_trend_2019 = sum(UII['2020'])/sum(TII[key])-1
                else:
                    # Si al final resulta 0 entonces le doy un 1 para que no me de el errores
                    # Despues tengo que pasarlo a porcentaje y me daria 100%. Esto significa que no existen datos
                    TII_trend_2019 = 1

# -------  Trend 2021 - 2020
            elif key == '2020':
                if sum(TII[key]) > 0:
                    TII_trend_2120 = sum(TII['2021'])/sum(TII[key])-1
                else:
                    TII_trend_2120 = 1

# -------  Trend 2022 - 2021
            elif key == '2021':
                if sum(TII[key]) > 0:
                    TII_trend_2221 = sum(TII['2022'])/sum(TII[key])-1
                else:
                    TII_trend_2221 = 1

            else:
                pass
# ----------------------------------------------------------------
#           CREO EL FRONT END
# ----------------------------------------------------------------

        # ------- Merge cells
        # Escojo el rango de cells
        # El primer numero es rows# Segundo numero es column
        # Utilizo la LETTER que cree para averiguar la ultima columan que debo de usar.
        myRange = Sheet3.Range(Sheet3.Cells((lastrow+2),1),Sheet3.Cells((lastrow+2),(lastcol-1)))



        myRange.MergeCells = True
        # Las options de align son : Center, Right,Left
        myRange.HorizontalAlignment = win32c.xlCenter
        # -------How to add colors
        #https://itecnote.com/tecnote/python-pywin32-excel-formatting/
        myRange.Interior.ColorIndex = 4
        #Shet3.Range("A1", "D1").Interior.ColorIndex = 6

        myRange.Value = 'Trend'
        Sheet3.Cells((lastrow+2),1).Font.Size = 14
        Sheet3.Cells((lastrow+2),1).Font.Bold = True
        Sheet3.Cells((lastrow+2),1).Font.Color = Book.RGB(0,0,0)
        Sheet3.Cells((lastrow+2),1).HorizontalAlignment = win32c.xlCenter

        #Sheet3.Range('A20','A26').Value =list
        # Yellow 6
        # Black 1
        # Yellow 2
        # Red 3
        # Green 4
        # Blue 5
        # Violet 7
        # Turquesa 8
        # Brown+red 9
        # Dark green 10
        # Dark blue 11
        # OCRE 12
        # Dark Purple 13

        # Marco todas las casillas de trend en verde
        #Sheet3.Range('A'+str(lastrow+4),letter+str(lastrow+4)).Interior.ColorIndex = 4

        #Sheet3.Range('A'+str(lastrow+4),letter+str(lastcol)).Interior.ColorIndex = 6
        # ----------------------------------------------------------------
        #           EDGES
        # ----------------------------------------------------------------
        # Aqui le estamos dando formato al recudro con el titulo
        #https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
        myRange.Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        myRange.Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        myRange.Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        myRange.Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)

        myRange.Borders(win32c.xlInsideHorizontal).Color = Book.RGB(0, 0, 0)
        myRange.Borders(win32c.xlInsideVertical).Color = Book.RGB(0, 0, 0)



        all_2019 = [sessions_trend_2019, TIR_trend_2019, UIR_trend_2019, TII_trend_2019, UII_trend_2019]
        all_2120 = [sessions_trend_2120, TIR_trend_2120, UIR_trend_2120, TII_trend_2120, UII_trend_2120]
        all_2221 = [sessions_trend_2221, TIR_trend_2221, UIR_trend_2221, TII_trend_2221, UII_trend_2221]

#---------------------------------------------------------------------------------------------
        for i in all_2019:
            if i == 1:
                pass

            else:
                Sheet3.Cells((lastrow+3),1).Value = '2019 vs 2020'
                Sheet3.Cells((lastrow+3),1).Font.Bold = True
                Sheet3.Cells((lastrow+3),1).Font.Size = 12


        for i in all_2120:
            if i == 1:
                pass
            else:
                Sheet3.Cells((lastrow+4),1).Value = '2020 vs 2021'
                Sheet3.Cells((lastrow+4),1).Font.Bold = True
                Sheet3.Cells((lastrow+4),1).Font.Size = 12


        for i in all_2221:
            if i == 1:
                pass
            else:
                Sheet3.Cells((lastrow+5),1).Value = '2022 vs 2021'
                Sheet3.Cells((lastrow+5),1).Font.Bold = True
                Sheet3.Cells((lastrow+5),1).Font.Size = 12


#-----------------------------------------------
#------------   SESSIONS  ----------------------
#-----------------------------------------------
# Aqui tengo que decir que es menor a 1 porque el resultado me lo da en FLOAT y no numeros enteros
# Poreso despues lo tengo que pasar a PORCENTAGE
        if  self.session_select and (sessions_trend_2019 < 1):

            col = 2
            row =3
            Sheet3.Cells((lastrow+row),1).Value = '2019 vs 2020'
            Sheet3.Cells((lastrow+row),col).Value = sessions_trend_2019

            if sessions_trend_2019 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass

        #-----------------------------------------------
        if  self.session_select and (sessions_trend_2120 < 1):

            col = 2
            row =4
            Sheet3.Cells((lastrow+row),1).Value = '2020 vs 2021'
            Sheet3.Cells((lastrow+row),col).Value = sessions_trend_2120

            if sessions_trend_2120 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)

        else:
            pass
            #-----------------------------------------------
        if  self.session_select and (sessions_trend_2221 <1):

            col = 2
            row =5
            Sheet3.Cells((lastrow+row),1).Value = '2021 vs 2022'
            Sheet3.Cells((lastrow+row),col).Value = sessions_trend_2221

            if sessions_trend_2221 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
#-----------------------------------------------
#------------   TII       ----------------------
#----------------------------------------------
        if  self.TII_select  and ( TII_trend_2019 < 1):

            col = 3
            row =3
            Sheet3.Cells((lastrow+row),col).Value = TII_trend_2019

            if TII_trend_2019 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #-----------------------------
        if  self.TII_select and ( TII_trend_2120 < 1):

            col = 3
            row =4
            Sheet3.Cells((lastrow+row),col).Value = TII_trend_2120

            if TII_trend_2120 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #-----------------------------
        if  self.TII_select and  ( TII_trend_2221 < 1):

            col = 3
            row =5
            Sheet3.Cells((lastrow+row),col).Value = TII_trend_2221

            if TII_trend_2221 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass

#-----------------------------------------------
#------------   UII       ----------------------
#-----------------------------------------------

        # Aqui jugamos con el last column para que solo me aparezca si hay un cirto numero de columnas.
        if  self.UII_select == 'Unique_Item_Investigations' and ( UII_trend_2019 < 1):

            col = 4
            row =3
            Sheet3.Cells((lastrow+row),col).Value = UII_trend_2019

            if UII_trend_2019 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #-------------------------------------I
        if  self.UII_select == 'Unique_Item_Investigations' and ( UII_trend_2120 < 1):

            col = 4
            row =4
            Sheet3.Cells((lastrow+row),col).Value = UII_trend_2120

            if UII_trend_2120 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #----------------------------------------------------
        if  self.UII_select == 'Unique_Item_Investigations'  and ( UII_trend_2221 < 1):
            col = 4
            row =5
            Sheet3.Cells((lastrow+row),col).Value = UII_trend_2221

            if UII_trend_2221 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass

#-----------------------------------------------
#------------   TIR       ----------------------
#----------------------------------------------
        if  self.TIR_select == 'Total_Item_Requests' and ( TIR_trend_2019 < 1):

            col = 5
            row =3
            Sheet3.Cells((lastrow+row),col).Value = TIR_trend_2019

            if TIR_trend_2019 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #-----------------------------
        if  self.TIR_select == 'Total_Item_Requests' and ( TIR_trend_2120 < 1):

            col = 5
            row =4
            Sheet3.Cells((lastrow+row),col).Value = TIR_trend_2120

            if TIR_trend_2120 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #-----------------------------
        if  self.TIR_select == 'Total_Item_Requests' and ( TIR_trend_2221 < 1):

            col = 5
            row =5
            Sheet3.Cells((lastrow+row),col).Value = TIR_trend_2221

            if TIR_trend_2221 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass


#-----------------------------------------------
#------------   UIR       ----------------------
#-----------------------------------------------

        # Aqui jugamos con el last column para que solo me aparezca si hay un cirto numero de columnas.
        if  self.UIR_select == 'Unique_Item_Requests' and ( UIR_trend_2019 < 1):

            col = 6
            row =3
            Sheet3.Cells((lastrow+row),col).Value = UIR_trend_2019

            if UIR_trend_2019 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #--------------------------------------
        if  self.UIR_select == 'Unique_Item_Requests' and ( UIR_trend_2120 < 1):

            col = 6
            row =4
            Sheet3.Cells((lastrow+row),col).Value = UIR_trend_2120

            if UIR_trend_2120 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass
            #----------------------------------------------------
        if  self.UIR_select == 'Unique_Item_Requests'  and ( UIR_trend_2221 < 1):
            col = 6
            row =5
            Sheet3.Cells((lastrow+row),col).Value = UIR_trend_2221

            if UIR_trend_2221 > 0:
                rgb = [15,209,60]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
            else:
                rgb = [0,0,255]
                Book.style_trend(self, Sheet3,rgb,row,col,lastrow,letter)
        else:
            pass

    def style_trend(self, Sheet3,rgb,row,col, lastrow,letter):

        Sheet3.Cells((lastrow+row),col).Font.Bold = True
        Sheet3.Cells((lastrow+row),col).Font.Color = Book.RGB(rgb[0],rgb[1],rgb[2])

        # ----------------------------------------------------------------
        #           EDGES
        # ----------------------------------------------------------------
        #https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
        # Esto es para que pueda hacer el GRID de la primera columan donde salen las fechas
        Sheet3.Cells((lastrow+row),1).Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),1).Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),1).Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),1).Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)


        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlEdgeBottom).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlEdgeLeft).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlEdgeRight).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlEdgeTop).Color = Book.RGB(0, 0, 0)

        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlInsideHorizontal).Color = Book.RGB(0, 0, 0)
        Sheet3.Cells((lastrow+row),col).Borders(win32c.xlInsideVertical).Color = Book.RGB(0, 0, 0)

        if col > 1:
            Sheet3.Cells((lastrow+row),col).Style = "Percent"
            Sheet3.Cells(row,col).Font.Size = 13
            Sheet3.Cells(row,col).HorizontalAlignment = win32c.xlCenter
        else:
            pass

    def webGraph(self):

        file = self.TEMP_SAVE_PATH + 'step_2.xlsx'
        self.df = pd.read_excel(file)

        count_years = self.df.Year.unique()
        count_months = self.df['Month-Year'].unique()
    #-------------------------------------
    #---------- YEAR - --------------#
    #-------------------------------------
        #-----------------------------------
        sum_session_yr = self.df.groupby('Year')['Sessions'].sum()
        sum_session_yr.values.tolist()
        sessions = self.df['Sessions'].sum()

        #-----------------------------------
        sum_TII_yr = self.df.groupby('Year')['Total_Item_Investigations'].sum()
        sum_TII_yr.values.tolist()
        TII = self.df['Total_Item_Investigations'].sum()

        #-----------------------------------
        sum_TIR_yr = self.df.groupby('Year')['Total_Item_Requests'].sum()
        sum_TIR_yr.values.tolist()
        TIR = self.df['Total_Item_Requests'].sum()

        #-----------------------------------
        sum_UII_yr = self.df.groupby('Year')['Unique_Item_Investigations'].sum()
        sum_UII_yr.values.tolist()
        UII = self.df['Unique_Item_Investigations'].sum()

        #--------  ---------------------------
        sum_UIR_yr = self.df.groupby('Year')['Unique_Item_Requests'].sum()
        sum_UIR_yr.values.tolist()
        UIR = self.df['Unique_Item_Requests'].sum()


    #-------------------------------------
    #---------- Monntly - --------------#
    #-------------------------------------
        #-----------------------------------
        sum_session_month = self.df.groupby('Month-Year')['Sessions'].sum()
        sum_session_month.values.tolist()
        sessions_m = self.df['Sessions'].sum()

        #-----------------------------------
        sum_TII_month = self.df.groupby('Month-Year')['Total_Item_Investigations'].sum()
        sum_TII_month.values.tolist()
        TII_m = self.df['Total_Item_Investigations'].sum()

        #-----------------------------------
        sum_TIR_month = self.df.groupby('Month-Year')['Total_Item_Requests'].sum()
        sum_TIR_month.values.tolist()
        TIR_m = self.df['Total_Item_Requests'].sum()

        #-----------------------------------
        sum_UII_month = self.df.groupby('Month-Year')['Unique_Item_Investigations'].sum()
        sum_UII_month.values.tolist()
        UII_m = self.df['Unique_Item_Investigations'].sum()

        #--------  ---------------------------
        sum_UIR_month = self.df.groupby('Month-Year')['Unique_Item_Requests'].sum()
        sum_UIR_month.values.tolist()
        UIR_m = self.df['Unique_Item_Requests'].sum()


        #-------------------------------------
        #---------- Monntly - --------------#
        #-------------------------------------
        #-----------------------------------

        if self.Yearly_Comparison == 'Yearly_Comparison':
            average_yr = final
        else:
            average_yr = ['.']


        return average_yr, count_years,count_months,sum_session_yr ,sum_TIR_yr,sum_TII_yr,sum_UIR_yr,sum_UII_yr,sessions,TIR,TII,UIR,UII,sum_session_month ,sum_TIR_month,sum_TII_month,sum_UIR_month,sum_UII_month,sessions_m,TIR_m,TII_m,UIR_m,UII_m

    def counter(self,KEY):

        #goodNum = open("counter_reports.txt",'r')

        f = open("apps/counter_webstats.txt", "a")
        f.write('1\n')
        f.close()
